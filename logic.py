import json
import re
import openpyxl
from llm import *
from input import *
from openpyxl.styles import PatternFill

def process_excel(file_path, description):
    """
    Process excel file, send data to gpt4 and get response, save to excel file
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    features = []
    for row in range(2, sheet.max_row + 1):
        feature = sheet.cell(row=row, column=1).value
        # Skip lines starting with '//'
        if feature and not str(feature).startswith('//'):
            features.append(f"\"{feature}\": \"\"")

    combined_prompt = get_command_prompt(description, '{' + ', '.join(features) + '}')

    print(combined_prompt)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    json_response = get_gpt_response(combined_prompt)

    print(json_response)

    # Find and parse json
    try:
        if json_response.startswith('{') and json_response.endswith('}'):
            responses = json.loads(json_response)
        else:
            json_match = re.search(r'```json\s*{(.+?)}\s*```', json_response, re.DOTALL)
            if json_match:
                json_extract = "{" + json_match.group(1) + "}"
                responses = json.loads(json_extract)
            else:
                raise ValueError("Nie znaleziono odpowiedzi JSON w odpowiedzi API")

    except (json.JSONDecodeError, ValueError) as e:
        print("Błąd w przetwarzaniu JSON: ", e)
        print("Odpowiedź: ", json_response)
        return

    # Update the sheet, add in new column
    newColumn = sheet.max_column + 1
    for row in range(2, sheet.max_row + 1):
        feature = sheet.cell(row=row, column=1).value
        if feature and feature in responses:
            response_column = newColumn
            cell = sheet.cell(row=row, column=response_column)
            cell.value = responses[feature]

            if cell.value == "???":
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.fill = yellow_fill

    while True:
        try:
            wb.save(file_path)
            break
        except Exception as e:
            print(f"Błąd podczas zapisu pliku: {e}")
            input("Upewnij się, że plik nie jest otwarty w innym programie, a następnie naciśnij Enter, aby spróbować ponownie.")

    print("Success!")